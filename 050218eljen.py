#OWTS DESIGN ELJEN
#needs pyexcel

############ADJUST CALC PER FILL ################################

#COPYRIGHT NARRAGANSETT ENGINEERING INC
#NEI-CDS.COM

#WORK PRODUCT. NOT FOR DISTRIBUTION
#program in tank size.
#print to word.
#include sketch

#USE FUNCTIONS!!!!

"""def bouy (weight, area):
    buyreq = weight * area
    print ('1000 tank :' , buyreq)
    return buyreq

buyreq = bouy ( 1000, 50)
print ('1500 tank:' , buyreq)


buyreq = bouy ( 1100, 40)
print ('1600 tank:' , buyreq)
"""




#load modules. make sure pip install



import time
from time import sleep
import datetime
from PIL import Image
import math


import argparse
import re
import os

import pathlib

import re
import os

import json
import datetime
from copy import deepcopy


import json

import datetime

from openpyxl import Workbook
import openpyxl
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import coordinate_from_string

wb = Workbook()

ws = wb.active

now = datetime.datetime.now()


#COL WIDTHS


dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
for col, value in dims.items():
    ws.column_dimensions[col].width = value



#site info

print ('FLAT, GRAVITY ELJEN DESIGN. WORKS BEST FOR NEW CONSTRUCTION SITES')
print ('NEEDS EDITS IF PUMP IS USED, OR MORE THAN 4 BEDROOMS')


print ('Please have all existing grades, watertable, ledge, and soils information ready')

projectno = input("Enter Project No: ")
ws['G2'] = projectno
ws['F2'] = ('Project Number (NEI): ')

address = input("Enter Address: ")
ws['G3'] = address
ws['F3'] = ('Site Address')


ws['J2'] = ('Not Valid for sloping fields')
ws['J3'] = ('Not Valid for pumped systems')
ws['J4'] = ('All Units in Decimal Feet U/N/O')
ws['J5'] = ('Contractor: Please Call out Exclusions')
ws['J5'] = ('Contractor: Please Call out Addt\'l Excav. Unit Costs')

#no br and flow

br = int(input("Enter number of bedrooms: (whole no,e.g. 2, only): "))
tflow =  br * 115

print ('total daily flow = ')
print (tflow)

ws['F5'] = ('Design Information')

ws['G6'] = tflow
ws['F6'] = ('Total Daily Flow, Gallons per day (GPD)')




#LOAD RATE

print ('enter soil cat, example:  1 , or 6m   (number and modifier no spaces: ')
print ('input HIGHEST (e.g. worst) withn 5\' of system;')
soilcat = input('soil category ex 1 or 1m (no spaces): ')


ws['G8'] = soilcat
ws['F8'] = ('Soil Category')

load61 = ["1m", "2", "4", "6", "7m"]   #loading rate groups
load70 = ["1", "3", "6m"]
load52 = ["5", "7"]
load48 = ["8m"]
load46 = ["8"]
load43 = ["9m"]
load40 = ["9"]



if soilcat in load61:
    loadrate = 0.61
    print (loadrate)
elif soilcat in load70:
    loadrate = 0.70
elif soilcat in load52:
    loadrate = 0.52
    print (loadrate)
elif soilcat in load48:
    loadrate = 0.48
    print (loadrate)
elif soilcat in load46:
    loadrate = 0.46
    print (loadrate)
elif soilcat in load43:
    loadrate = 0.43
    print (loadrate)
elif soilcat in load40:
    loadrate = 0.40
    print (loadrate)
else:
    print ('improper soil category')

ws['G9'] = loadrate
ws['F9'] = ('Loading Rate')


sizereqd = tflow/loadrate     #min size reqd per br flow / loadrate
print ('system size required: ')
print("%.2f" % sizereqd)

ws['G10'] = sizereqd
ws['F10'] = ('Min. Leach Area')
ws['H10'] = ('(Flow / Loadrate)')


elj = sizereqd / 28

minelj = (math.ceil(elj))   #min no eljens (rounded up from size require)
print(minelj)   
#print(math.ceil(minelj))

ws['G11'] = minelj
ws['F11'] = ('Min. No. Eljens') 
ws['H11'] = ('Min. Leaching area / 28) rounded up')


sizepvd = minelj * 28
ws['G12'] = sizepvd
ws['F12'] = ('Leaching Area Provided')

    #=disc/loadrate




print ('are you in a critical resource area (see regs,e.g. near salt pond area Y/N ')   #ADD CRITICAL RESOURCE AREA IN LATER
#crit = input
print ('critical resource area values not programmed in yet')


print ('are you in a nitrogen sensitive area ? Y/N ')   #ADD NITROGEN AREA IN LATER
#crit = input
print ('nitrogen values not programmed in yet')

print ('NOT VALID FOR SLOPING SITES')

print ('If there is fill, adjustment is needed; adjust highest or mean grade if fill present')
hwt = int(input("enter HIGHEST (aka worst) SHGWT depth in area system IN INCHES (eg:  60):  INCLUDE FILL")) #edit fill ____________________________
fill = int(input("enter amount of fill in inches: , enter 0\" if no fill present")) ##050218


############ADJUST CALC PER FILL



hwtft = hwt / 12   #wwater table in feet - ACTUAL DEPTH FROM GRADE

hwtftX = hwtft - fill   #HTWFTX = WATER TABLLE DEPTH - FILL DEPTH TO YIELD TRUE WATER TABLE

print ('soil eval in ft')
print (hwtft)

ledge = int(input('enter HIGHEST (aka worst) ledge elevation IN INCHES (eg: 100): '))

ledgeft = ledge / 12
print ('ledge in ft')
print (ledgeft)


#not working need to code in range
if hwtftX < 2:
    print ('DO NOT USE ELJENS - ADVANCED TREATMENT REQUIRED')
    print ('I AM NOW CRASHING ON PURPOSE')
    ws['J4'] = ('DO NOT USE ELJENS - ADVANCED TREATMENT REQUIRED WT < 2 FT ')
    
    
elif hwtftX <= 4:                      #highest or mean system if water table > 4 ft or < 4 ft
    eg = int(input('enter HIGHEST elevation in area of system'))
    gledge = eg - ledgeft + 5
    ghwt = eg - hwtft + 3
    avghigh = 'highest existing grade'
    wtel = eg - hwtft  #wt elev = exist grade - wt in feet
    ledgeel = eg - ledgeft #ledge elev = exist grade - ledge in feet
    if ghwt >= gledge:
       g = ghwt + 0
       print ('eljen bottom; water table is limiting factor')
       print ("%.2f" % g)
       print ('SHGWT elevation: ')
       print ("%.2f" % ghwt)
       print ('ledge elevation: ')
       print ("%.2f" % gledge)
    elif gledge > ghwt:
       g = gledge
       print ('eljen bottom;  water table is limiting factor')
       print ("%.2f" % g)
       print ('SHGWT elevation: ')
       print ("%.2f" % ghwt)
       print ('ledge elevation: ')
       print ("%.2f" % gledge)
        
elif hwtftX > 4:
    eg = int(input('enter MEAN elevation in area of system'))
    gledge = eg -ledgeft + 5
    ghwt = eg - hwtft + 3
    avghigh = 'mean existing grade'
    wtel = eg - hwtft  #wt elev = exist grade - wt in feet
    ledgeel = eg - ledgeft #ledge elev = exist grade - ledge in feet
    if ghwt >= gledge:
       g = ghwt + 0
       print ('eljen bottom; water table is limiting factor')
       print ("%.2f" % g)
       print ('SHGWT elevation: ')
       print ("%.2f" % ghwt)
       print ('ledge elevation: ')
       print ("%.2f" % gledge)
       
       
    elif gledge > ghwt:
       g = gledge
       print ('eljen bottom;  water table is limiting factor')
       print ("%.2f" % g)
       print ('SHGWT elevation: ')
       print ("%.2f" % ghwt)
       print ('ledge elevation: ')
       print ("%.2f" % gledge)



#####add section, if maxgr < mean or highest grade then mean or highest grade becomes max gr.
    

        
else:
    print ('please enter numbers only for eval result')

#calcs for eljen

strip = g - 0.5
#strip = h - 0.5
inv = g + 0.58
maxgr = inv + 2.5
mingr = inv + 1.5
   

#printcalcs eljen

print ("%.2f" % maxgr)
print (' max grade over system')

print ("%.2f" % mingr)
print (' min grade over system')

print ("%.2f" % inv)
print (' eljen invert elevation')

#print ("%.2f" % h)  #bad math remove
#print (' bottom of ELJEN')

print ("%.2f" % g)
print(' eljen bottom')

print ("%.2f" % strip)
print ('strip elevation, 0.5\' depth, subject to bottom inspection. contractor to note additional excvation may be needed in contract')

print ("%.2f" % ghwt)
print ('SHGWT Elevation: ')

print ("%.2f" % gledge)    
print ('Ledge Elevation: ')

#slope calcs

#-------------------- move to end
##get these into excel sheet



########GO BACKWARDS FROM U TO G (E.G. FIELD TO HOUSE) ------------------------------


#Vvvv new
lfdbox = float(input("Enter line lenght from leachfield to dbox  "))
ws['U48'] = lfdbox
ws['U49'] = ('lenght tank to dbox.')

lfdboxslope = float(input("Enter line slope from leachfield to to dbox e.g. 0.01 to 0.03:  "))
ws['V48'] = lfdboxslope
ws['V49'] = ('slope tank to dbox.')


#U BLANK

#X

ws['W49'] = inv     #INVERT
ws['W48'] = ('Eljen Invert')
ws['W50'] = ('Verify matches in list')

#T

ws['T48'] = ('0.16')
ws['T49'] = ('dbox elev drop (2")')

#S

dboxout = inv + (lfdboxslope * lfdbox)
print (dboxout)
ws['S48'] = dboxout
ws['S49'] = ('dbox outlet invert')

#R
dboxin = dboxout + 0.16
print (dboxin)
ws['R48'] = dboxin
ws['R49'] = ('dbox inlet invert')





#O
tankbox = float(input("Enter line lenght from septic tank to dbox:  "))
ws['O48'] = tankbox
ws['O49'] = ('lenght tank to dbox.')

#P
tankboxboxslope = float(input("Enter line slope from septic tank to dbox e.g. 0.01 to 0.03:  "))
ws['P48'] = tankboxboxslope
ws['P49'] = ('slope tank to dbox.')





#Q   - AFTER P 0 
ws['Q48'] = tankboxboxslope * tankbox
ws['Q49'] = ('elev. drop ')

#N
             
tankout = dboxin + (tankbox * tankboxboxslope)
print (tankout)
ws['N48'] = tankout
ws['N49'] = ('tank outlet elev.')


#M

ws['M48'] = ('0.25\'')
ws['M49'] = ('elev drop in tank (3").')



#L


tankin = tankout + 0.25
print (tankin)
ws['L48'] = tankin
ws['L49'] = ('tank inv.')



#J
housetank = float(input("Enter line lenght from house to tank (25' min if possible) "))
ws['J48'] = housetank
ws['J49'] = ('slope house to tank ')


#i
housetankslope = float(input("Enter line slope e.g. 0.01: "))
ws['I48'] = housetankslope
ws['I49'] = ('slope house to tank ')

#K  AFTER J I 

diffht = housetank * housetankslope  #elev drop house to tank
ws['K48'] = diffht
ws['K49'] = ('elev. drop')



#H
housein = tankin + (housetank * housetankslope)
print (housein)
ws['H48'] = housein
ws['H49'] = ('Inv. at Structure ')


#g
fghouse = float(input("Enter finish grade at house invert: "))
ws['G48'] = fghouse
ws['G49'] = ('Min. Finish Grade at Structure')



if (housein + 1.25) > fghouse:
    print ('error, additional fill needed at invert to structure!')
    ws['G51'] = ('error, additional fill needed at invert to structure!')
else:
    print ('sufficient cover')
    ws['G51'] = ('sufficient cover over invert')
             
#tank size
#invert, bottom and top of tank
#buoyancy calcs

print ('Septic Tank and Buoyancy')

egtank = float(input("Enter Existing Grade at Septic Tank: "))
ws['E70'] = egtank         
ws['E71'] = ('Existing Grade at Septic Tank')

wttank = egtank - hwtft
ws['E73'] = wttank         
ws['E74'] = ('SHGWT El. at Septic Tank')

ledgetank = egtank - ledgeft    
ws['E76'] = ("%.2f" % ledgetank)         
ws['E77'] = ('Ledge El. at Septic Tank')
             

if br == 1:
    print ('deed restriction required')
    ws['L80'] = ('Deed Restriction Required!')
elif br == 2:
    print ('use 1000 gallon tank')
    ws['L80'] = ('Deed Restriction Required!')
    ws = wb.active
    my_png = openpyxl.drawing.image.Image('1000ST.png')
    ws.add_image(my_png, 'H60')   # tank location
    bottank = tankin - 4.58       #4.58'  (4' - 7")
    ws['R80'] = bottank
    ws['R79'] = ('Bottom of Tank')  
    toptank = bottank + 5.67        #top tank + 5'-8"       
    ws['R73'] = toptank
    ws['R72'] = ('Top of Tank')
    fgtank = toptank + 1.0 #1 min cover over tank
    ws['R70'] = fgtank         
    ws['R69'] = ('Finish Grade at Tank, Minimum')
    ws['R75'] = ('Min. 1 ft. cover over tank')
    buoyreqd1000 = (egtank - wttank) * (5.16 * 8.0) * 110    #grade - water table x tank dims x 110 cuft water
    ws['R87'] = buoyreqd1000        
    ws['R87'] = ('Buoyancy Required = ((ex. grade - water table) x tank dims x 110 #/cuft)')
    if buoyreqd1000 > 9000:
        weightreq = buoyreqd1000 - 9000   #9000 tank weight
        ftconc = weightreq / 6192    # 5.16 * 8 *150
        ws['R86'] = ('9,000 LBS')
        ws['S86'] = ('Tank Weight')
        ws['R89'] = "%.2f" % weightreq    #-----------------------------------
        ws['S89'] = ('LBS')
        ws['S90'] = ('Anti Float Base Required, tank dims x 150 psf conc')
        ws['R92'] = ("%.2f" % ftconc)    #-----------------------------------
        ws['S93'] = ('Depth of Anti Float Base (150 psf conc)')
    else:
        ws['R89'] = ('No Antifloation Base Required')
        ws['R86'] = ('9,000 LBS')
        ws['S86'] = ('Tank Weight')

             
elif br == 3 or 4:
    print ('use 1500 gallon tank')
    ws = wb.active
    my_png = openpyxl.drawing.image.Image('1500ST.png')
    ws.add_image(my_png, 'H60')   # tank location
    bottank = tankin - 4.58       #4.58'  (4' - 7") ok
    ws['R80'] = bottank
    ws['R79'] = ('Bottom of Tank')  
    toptank = bottank + 5.67        #top tank + 5'-8"       ok  
    ws['R73'] = toptank
    ws['R72'] = ('Top of Tank')
    fgtank = toptank + 1.0 #1 min cover over tank  ok
    ws['R70'] = fgtank         
    ws['R69'] = ('Finish Grade at Tank, Minimum')
    ws['R75'] = ('Min. 1 ft. cover over tank')
    buoyreqd1500 = (egtank - wttank) * (10.5 * 5.67) * 110    #grade - water table x tank dims x 110 cuft water
    ws['R87'] = buoyreqd1500        
    ws['S87'] = ('Buoyancy Required = ((ex. grade - water table) x tank dims x 110 #/cuft)')
    if buoyreqd1500 > 12930:
        weightreq = buoyreqd1500 - 12930   #12930 tank weight
        ftconc = weightreq / 8925    # 59.5 sqft tank *150
        ws['R86'] = ('12,930 LBS')
        ws['S86'] = ('Tank Weight')
        ws['R89'] = "%.2f" % weightreq    #-----------------------------------
        ws['S89'] = ('LBS')
        ws['S90'] = ('Anti Float Base Required, tank dims x 150 psf conc')
        ws['R92'] = "%.2f" % ftconc    #-----------------------------------
        ws['S93'] = ('Depth of Anti Float Base (150 psf conc)')
    else:
        ws['R89'] = ('No Antifloation Base Required')
        ws['R86'] = ('12,930 LBS')
        ws['S86'] = ('Tank Weight')

                
else:
    print ('determine tank size (1000 gal per 3 br and 250 per ea. additional')
    ws['M70'] = ('determine tank size (1000 gal per 3 br and 250 per ea. additional)')

if gledge > ledgetank:
    print ('ledge removal required!!')
    ws['E84'] = ('ledge removal required!!!')
else:
    print ('ledge ok')
        


#-------------------- move to end


    #START AT G12


    """wtel = eg - hwtft  #wt elev = exist grade - wt in feet
    ledgeel = eg - ledgeft #ledge elev = exist grade - ledge in feet"""



ws['G20'] = wtel
ws['F20'] = 'SHGWT El:'

ws['G25'] = ghwt
ws['F25'] = 'Min. Bottom El per water table: SHGWT + 3'


ws['G19'] = ledgeel
ws['F19'] = 'Ledge El:'

ws['G24'] = ("%.2f" % gledge)   # ERROR BY 5'
ws['F24'] = ('Min. Bottom El per water table: SHGWT + 3')

ws['G15'] = eg      #EX GRADE DO NOT USE (PRE HEIGHT DETERMINATION)
ws['F15'] = avghigh #PRINTS AVG OR MEAN


ws['G13'] = maxgr
ws['F13'] = ' Max Finish Grade'

ws['G14'] = mingr
ws['F14'] = ' Min Finish Grade'

ws['G16'] = inv     #INVERT
ws['F16'] = ' Eljen Invert'

ws['G17'] = g      #ELJEN BOT
ws['F17'] = ' Eljen Bottom'

ws['G18'] = strip    #BOT EXCV STRIP
ws['F18'] = ('Bottom of Sand, Minimum')
ws['H18'] = ('subject to bottom inspection')
ws['I18'] = ('by designer of record')

ws['G22'] = hwtft    #wt depth
ws['F22'] = 'Seasonal High Water Table Depth from EXISTING grade'
ws['H22'] = 'FT'

ws['G23'] = ("%.2f" % ledgeft)    #ledge depth
ws['F23'] = 'Ledge Depth from EXISTING grade'
ws['H23'] = 'FT'

ws['G24'] = hwtftX    #WATER TABLE - FILL (ACTUAL WATER TABLE)
ws['F24'] = 'Seasonal High Water Table Depth from original grade'
ws['H24'] = 'FT'          
                 


ws['G46'] = 'SECTION NO'
ws['H46'] = '1'
ws['L46'] = '2'
ws['N46'] = '3'
ws['R46'] = '4'
ws['S46'] = '5'
ws['X46'] = '6'

                 
ws = wb.active
my_png = openpyxl.drawing.image.Image('DBOX.png')
ws.add_image(my_png, 'G27')

ws = wb.active
my_png = openpyxl.drawing.image.Image('a1800.png')
ws.add_image(my_png, 'S1')

ws = wb.active
my_png = openpyxl.drawing.image.Image('eljen xc.png')
ws.add_image(my_png, 'N1')




wb.save("eljen-out.xlsx")
print ('done eljen-out xlsx')


    

